import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Load the Excel file
file_path = "45 Days placement-Absentees List Date wise.xlsx"
excel_file = pd.ExcelFile(file_path)

# Sheets and readable date map
sheet_names = excel_file.sheet_names[1:]  # Skip first sheet
date_map = {
    '1062025': '10/06/2025', '11625': '11/06/2025', '12625': '12/06/2025', '13625': '13/06/2025',
    '14625': '14/06/2025', '16062025': '16/06/2025', '17062025': '17/06/2025', '18062025': '18/06/2025',
    '19062025': '19/06/2025', '20062025': '20/06/2025', '21062025': '21/06/2025', '23062025': '23/06/2025',
    '24062025': '24/06/2025', '25062025': '25/06/2025', '26062025': '26/06/2025', '27062025': '27/06/2025',
    '28062025': '28/06/2025'
}

# Extract absentees from AIA class
def extract_aia_roll_numbers(df):
    aia_rolls = []
    for val in df.values.flatten():
        if isinstance(val, str) and 'AIA' in val.upper():
            aia_rolls.append(val.strip().upper())
    return aia_rolls

absentees_by_date = {}
for sheet in sheet_names:
    df = excel_file.parse(sheet, header=None)
    absentees_by_date[sheet] = extract_aia_roll_numbers(df)

# Use roll numbers from template
roll_numbers = [
    "22AIA01", "22AIA02", "22AIA03", "22AIA04", "22AIA05", "22AIA06", "22AIA07", "22AIA08", "22AIA09",
    "22AIA10", "22AIA11", "22AIA12", "22AIA13", "22AIA14", "22AIA15", "22AIA16", "22AIA17", "22AIA18",
    "22AIA19", "22AIA20", "22AIA21", "22AIA22", "22AIA23", "22AIA24", "22AIA26", "22AIA27", "22AIA28",
    "22AIA29", "22AIA30", "22AIA31", "22AIA32", "22AIA33", "22AIA34", "22AIA35", "22AIA36", "22AIA37",
    "22AIA38", "22AIA39", "22AIA40", "22AIA41", "22AIA42", "22AIA43", "22AIA44", "22AIA45", "22AIA46",
    "22AIA47", "22AIA49", "22AIA50", "22AIA51", "22AIA52", "22AIA53", "22AIA54", "22AIA55", "22AIA56",
    "22AIA57", "22AIA58", "22AIA59", "22AIA60", "22AIA61", "22AIA62", "22AIA65", "22AIA66", "22AIA67",
    "22AIA68", "22AIA70"
]

# Create workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Attendance_June2025"

# Header rows
ws.cell(row=1, column=1).value = "Roll No"
periods = [str(i) for i in range(1, 9)]

# Ordered date columns
ordered_dates = [date_map[k] for k in sorted(date_map, key=lambda x: pd.to_datetime(date_map[x], dayfirst=True))]
col_index = 2
for date in ordered_dates:
    ws.cell(row=1, column=col_index).value = date
    for i in range(8):
        ws.cell(row=2, column=col_index + i).value = periods[i]
    col_index += 8

# Fill attendance data
for i, roll in enumerate(roll_numbers, start=3):
    ws.cell(row=i, column=1).value = roll
    col = 2
    for sheet_key in sorted(date_map, key=lambda x: pd.to_datetime(date_map[x], dayfirst=True)):
        absentees = absentees_by_date.get(sheet_key, [])
        for _ in range(8):
            status = "A" if roll.upper() in absentees else "P"
            ws.cell(row=i, column=col).value = status
            col += 1

# Save final sheet
output_file = "AIA_Attendance_June2025_Combined.xlsx"
wb.save(output_file)
print(f"✅ File created: {output_file}")
